from odoo import api, fields, models, _
from odoo.exceptions import UserError
from datetime import datetime, date
from io import BytesIO
import base64
from openpyxl import load_workbook


class ShiftAssignmentMatrixImportWizard(models.TransientModel):
    _name = 'shift.assignment.matrix.import.wizard'
    _description = 'Import Shift Assignment Matrix from Excel'

    matrix_id = fields.Many2one('shift.assignment.matrix', string=_('Matrix'), required=True, ondelete='cascade')
    file_data = fields.Binary(string=_('Excel File'), required=True, help=_('Upload an Excel file to import shifts'))
    file_name = fields.Char(string=_('File Name'))

    def _parse_header_date(self, value):
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)):
            # Avoid trying to parse Excel serials; require text dd/mm/YYYY
            return None
        try:
            return datetime.strptime(str(value).strip(), '%d/%m/%Y').date()
        except Exception:
            return None

    def _apply_import(self, matrix, ws):
        # Build header date map (row 1)
        column_index_to_date = {}
        max_col = ws.max_column or 1
        for col_idx in range(2, max_col + 1):  # start from column B (index 2)
            parsed = self._parse_header_date(ws.cell(row=1, column=col_idx).value)
            if parsed:
                column_index_to_date[col_idx] = parsed
        if not column_index_to_date:
            raise UserError(_('No valid date columns found in the first row (expected dd/mm/YYYY).'))

        # Determine date range from Excel and employees from first column
        excel_dates = sorted(column_index_to_date.values())
        start_date = excel_dates[0]
        end_date = excel_dates[-1]

        Employee = self.env['hr.employee']
        Calendar = self.env['resource.calendar']

        max_row = ws.max_row or 1
        employees_found = []
        missing_employees = []
        for row_idx in range(2, max_row + 1):
            name_cell = ws.cell(row=row_idx, column=1).value
            if not name_cell:
                continue
            employee = Employee.search([('name', '=', name_cell)], limit=1)
            if employee:
                employees_found.append(employee.id)
            else:
                missing_employees.append((row_idx, name_cell))

        # Update matrix with dates and employees from Excel
        vals = {
            'start_date': start_date,
            'end_date': end_date,
        }
        if employees_found:
            vals['employee_ids'] = [(6, 0, list(set(employees_found)))]
        # Set default name based on dates if empty
        if not matrix.name and start_date and end_date:
            try:
                vals['name'] = _('Shift %s to %s') % (start_date.strftime('%d/%m/%Y'), end_date.strftime('%d/%m/%Y'))
            except Exception:
                pass
        matrix.write(vals)

        # Ensure lines exist for the Excel-defined range/employees
        if not matrix.line_ids and matrix.start_date and matrix.end_date and matrix.employee_ids:
            matrix._prepare_lines()

        # Index by (employee_id, date)
        lines_by_employee_date = {}
        for line in matrix.line_ids:
            if not line.employee_id or not line.date:
                continue
            lines_by_employee_date[(line.employee_id.id, fields.Date.to_date(line.date))] = line

        errors = []
        for row_idx in range(2, max_row + 1):
            name_cell = ws.cell(row=row_idx, column=1).value
            if not name_cell:
                continue
            employee = Employee.search([('name', '=', name_cell)], limit=1)
            if not employee:
                errors.append(_('Row %(row)d: Employee "%(name)s" not found.') % {
                    'row': row_idx,
                    'name': name_cell,
                })
                continue
            for col_idx, date_value in column_index_to_date.items():
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value in (None, '', 0):
                    continue
                calendar = Calendar.search([('name', '=', str(cell_value).strip())], limit=1)
                if not calendar:
                    errors.append(_('Row %(row)d, Column %(col)d (%(date)s): Shift "%(shift)s" not found in Resource Calendars.') % {
                        'row': row_idx,
                        'col': col_idx,
                        'date': date_value.strftime('%d/%m/%Y'),
                        'shift': cell_value,
                    })
                    continue
                line = lines_by_employee_date.get((employee.id, date_value))
                if not line:
                    # If for any reason the line is missing, create it on the fly
                    line = self.env['shift.assignment.matrix.line'].create({
                        'matrix_id': matrix.id,
                        'employee_id': employee.id,
                        'date': date_value,
                    })
                    lines_by_employee_date[(employee.id, date_value)] = line
                line.calendar_id = calendar.id

        if missing_employees:
            for row_idx, name in missing_employees:
                errors.append(_('Row %(row)d: Employee "%(name)s" not found.') % {
                    'row': row_idx,
                    'name': name,
                })

        if errors:
            raise UserError(_('Import completed with errors:\n%s') % ('\n'.join(errors)))

    def action_import(self):
        self.ensure_one()
        if not self.file_data:
            raise UserError(_('Please upload an Excel file.'))
        if not self.matrix_id:
            raise UserError(_('Matrix is required.'))
        file_bytes = base64.b64decode(self.file_data)
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        ws = wb.active if wb.active else wb[wb.sheetnames[0]]
        self._apply_import(self.matrix_id, ws)
        return {
            'type': 'ir.actions.act_window',
            'name': _('Shift Assignment Matrix'),
            'res_model': 'shift.assignment.matrix',
            'view_mode': 'form',
            'res_id': self.matrix_id.id,
            'target': 'current',
        }


